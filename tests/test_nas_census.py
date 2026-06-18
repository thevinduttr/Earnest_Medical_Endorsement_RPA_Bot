from __future__ import annotations

from datetime import date
from pathlib import Path
import re
import tempfile
import unittest
from unittest.mock import patch
import warnings
from zipfile import ZipFile

import pandas as pd
from openpyxl import load_workbook

from src.services.census_service.nas.addition_census import (
    DEFAULT_NAS_ADD_TEMPLATE,
    build_nas_addition_census_file,
    value_from_nas_addition_row,
)
from src.services.census_service.nas.deletion_census import (
    DEFAULT_NAS_DELETE_TEMPLATE,
    build_nas_deletion_census_file,
    value_from_nas_deletion_row,
)
from src.services.census_service.sukoon.common import normalize_header


EXT_LIST_PATTERN = re.compile(
    br"<(?:\w+:)?extLst\b.*?</(?:\w+:)?extLst>",
    flags=re.DOTALL,
)


class NasCensusMappingTests(unittest.TestCase):
    def setUp(self) -> None:
        self.addition_row = {
            "ContractName": "Fw: JASHANMAL NATIONAL CO LLC",
            "FirstName": "Jane",
            "MiddleName": "Mary",
            "LastName": "Doe",
            "ArabicFirstName": "Arabic First",
            "ArabicMiddleName": "Arabic Middle",
            "ArabicLastName": "Arabic Last",
            "EffectiveDate": date(2026, 6, 18),
            "DateOfBirth": date(1990, 1, 2),
            "Gender": "Female",
            "MaritalStatus": "Single",
            "Category": "CAT A-DXB",
            "Relation": "Principal",
            "Department": "Operations",
            "Grade": "A",
            "PrincipalCardNo": "PC-1",
            "FamilyNo": "F-1",
            "StaffId": "S-1",
            "Nationality": "Sri Lanka",
            "SubNationality": "Other",
            "EmiratesId": "784-0000",
            "UnifiedNo": "U-1",
            "PassportNo": "P-1",
            "WorkCountry": "United Arab Emirates",
            "WorkEmirate": "Dubai",
            "WorkRegion": "Al Barsha",
            "ResidenceCountry": "United Arab Emirates",
            "ResidenceEmirate": "Dubai",
            "ResidenceRegion": "Jumeirah",
            "Email": "jane@example.com",
            "MobileNo": "0500000000",
            "SalaryBand": "Salary between 4,001 and 12,000 AED per month",
            "Commission": "Y",
            "VisaIssuanceEmirate": "Dubai",
            "BirthCertificateNumber": "BC-1",
            "VisaFileNumber": "VF-1",
            "MemberPhoto": "jane.jpg",
            "MemberType": "Expat whose residence issued in Dubai",
            "Occupation": "ACCOUNTANT",
            "RegulatorNo": "R-1",
            "PassportExpiryDate": date(2030, 12, 31),
        }

    def test_every_addition_header_is_mapped(self) -> None:
        workbook = load_workbook(DEFAULT_NAS_ADD_TEMPLATE, read_only=True)
        worksheet = workbook["Sample Template"]
        values = {
            normalize_header(cell.value): value_from_nas_addition_row(
                self.addition_row,
                normalize_header(cell.value),
                1,
            )
            for cell in worksheet[1]
            if normalize_header(cell.value)
        }
        workbook.close()

        self.assertEqual(49, len(values))
        self.assertEqual("JASHANMAL NATIONAL CO LLC", values["contract name"])
        self.assertEqual("2026-06-18", values["effective date"])
        self.assertEqual("1990-01-02", values["dob"])
        self.assertEqual("Yes", values["commission"])
        self.assertEqual("2030-12-31", values["passport expiry date"])
        for optional_header in (
            "visa expiry date",
            "coc available",
            "pec declaration",
            "medical declaration file",
            "visa type",
            "company phone",
            "company mail",
            "waived pec declaration",
        ):
            self.assertEqual("", values[optional_header])

    def test_deletion_mapping_and_card_fallback(self) -> None:
        row = {
            "HealthCardNumber": "",
            "StaffId": "STAFF-9",
            "EmiratesId": "784-9",
            "DeletionEffectiveDate": date(2026, 6, 18),
        }

        self.assertEqual(
            "STAFF-9",
            value_from_nas_deletion_row(row, "member card no", 1),
        )
        self.assertEqual(
            "18-06-2026",
            value_from_nas_deletion_row(row, "effective date", 1),
        )
        self.assertEqual(
            "",
            value_from_nas_deletion_row(row, "deletion reason", 1),
        )


class NasCensusWorkbookTests(unittest.TestCase):
    @staticmethod
    def _extension_xml(path: Path) -> bytes:
        with ZipFile(path) as archive:
            worksheet_xml = archive.read("xl/worksheets/sheet1.xml")
        match = EXT_LIST_PATTERN.search(worksheet_xml)
        if match is None:
            raise AssertionError(f"Worksheet extension XML missing from {path}")
        return match.group(0)

    def test_addition_workbook_preserves_template_and_writes_rows(self) -> None:
        records = pd.DataFrame(
            [
                {
                    "ContractName": "Fw: Contract One",
                    "FirstName": "First",
                    "LastName": "Member",
                    "EffectiveDate": date(2026, 6, 18),
                    "DateOfBirth": date(1990, 1, 2),
                    "Commission": "No",
                },
                {
                    "ContractName": "Contract Two",
                    "FirstName": "Second",
                    "LastName": "Member",
                    "EffectiveDate": date(2026, 6, 19),
                    "DateOfBirth": date(1991, 2, 3),
                    "Commission": "Yes",
                },
            ]
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "member_addition.xlsx"
            with patch(
                "src.services.census_service.nas.common.load_request_members_dataframe",
                return_value=records,
            ):
                result = build_nas_addition_census_file(
                    request_id="REQ-1",
                    output_path=output_path,
                    include_user_ids=["U1", "U2"],
                )

            self.assertEqual(2, result.members_count)
            self.assertEqual(
                self._extension_xml(DEFAULT_NAS_ADD_TEMPLATE),
                self._extension_xml(output_path),
            )

            with warnings.catch_warnings():
                warnings.simplefilter("ignore", UserWarning)
                template = load_workbook(DEFAULT_NAS_ADD_TEMPLATE, read_only=True)
                generated = load_workbook(output_path, read_only=True, data_only=False)

            self.assertEqual(template.sheetnames, generated.sheetnames)
            self.assertEqual(
                sorted(template.defined_names),
                sorted(generated.defined_names),
            )
            worksheet = generated["Sample Template"]
            self.assertEqual("Contract One", worksheet["A2"].value)
            self.assertEqual("First", worksheet["B2"].value)
            self.assertEqual("Contract Two", worksheet["A3"].value)
            self.assertEqual("General", worksheet["H2"].number_format)
            self.assertEqual("2026-06-18", worksheet["H2"].value)
            template.close()
            generated.close()

    def test_deletion_workbook_writes_general_date_text(self) -> None:
        records = pd.DataFrame(
            [
                {
                    "HealthCardNumber": "CARD-1",
                    "StaffId": "STAFF-1",
                    "EmiratesId": "784-1",
                    "DeletionEffectiveDate": date(2026, 6, 18),
                }
            ]
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "member_deletion.xlsx"
            with patch(
                "src.services.census_service.nas.common.load_request_members_dataframe",
                return_value=records,
            ):
                build_nas_deletion_census_file(
                    request_id="REQ-2",
                    output_path=output_path,
                    include_user_ids=["U1"],
                )

            self.assertEqual(
                self._extension_xml(DEFAULT_NAS_DELETE_TEMPLATE),
                self._extension_xml(output_path),
            )
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", UserWarning)
                generated = load_workbook(output_path, read_only=True, data_only=False)
            worksheet = generated["Sample Template"]
            self.assertEqual("CARD-1", worksheet["A2"].value)
            self.assertEqual("784-1", worksheet["B2"].value)
            self.assertEqual("18-06-2026", worksheet["C2"].value)
            self.assertEqual("General", worksheet["C2"].number_format)
            self.assertEqual("", worksheet["D2"].value or "")
            generated.close()


if __name__ == "__main__":
    unittest.main()
