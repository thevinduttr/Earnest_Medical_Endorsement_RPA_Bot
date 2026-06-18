from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
import os
import posixpath
import re
from pathlib import Path
from typing import Any, Callable, Dict, Iterable
from uuid import uuid4
import warnings
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook

from src.services.census_service.sukoon.common import (
    CensusBuildResult,
    load_request_members_dataframe,
    normalize_header,
)


ValueMapper = Callable[[Dict[str, Any], str, int], Any]

_WORKBOOK_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_ROOT_TAG_PATTERN = re.compile(br"<(?:\w+:)?worksheet\b[^>]*>")
_EXT_LIST_PATTERN = re.compile(
    br"<(?:\w+:)?extLst\b.*?</(?:\w+:)?extLst>",
    flags=re.DOTALL,
)


def normalize_scalar(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return format(value, "f").rstrip("0").rstrip(".") or "0"
    return str(value).strip()


def normalize_contract_name(value: Any) -> str:
    text = normalize_scalar(value)
    return re.sub(r"^(?:(?:FW|FWD|RE)\s*:\s*)+", "", text, flags=re.IGNORECASE).strip()


def normalize_commission(value: Any) -> str:
    if value is None or str(value).strip() == "":
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"

    normalized = str(value).strip().upper()
    if normalized in {"1", "Y", "YES", "TRUE", "T"}:
        return "Yes"
    if normalized in {"0", "N", "NO", "FALSE", "F"}:
        return "No"
    return str(value).strip()


def format_iso_date(value: Any) -> str:
    parsed = _coerce_date(value)
    return parsed.strftime("%Y-%m-%d") if parsed else normalize_scalar(value)


def format_day_first_date(value: Any) -> str:
    parsed = _coerce_date(value)
    return parsed.strftime("%d-%m-%Y") if parsed else normalize_scalar(value)


def _coerce_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value or "").strip()
    if not text:
        return None

    for date_format in (
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%m/%d/%Y",
    ):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue
    return None


def _worksheet_part_path(workbook_path: Path, sheet_name: str) -> str:
    from xml.etree import ElementTree

    with ZipFile(workbook_path) as archive:
        workbook_xml = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        relationships_xml = ElementTree.fromstring(
            archive.read("xl/_rels/workbook.xml.rels")
        )

    relationship_id = None
    for sheet in workbook_xml.findall(f".//{{{_WORKBOOK_NS}}}sheet"):
        if sheet.get("name") == sheet_name:
            relationship_id = sheet.get(f"{{{_REL_NS}}}id")
            break
    if not relationship_id:
        raise ValueError(f"Worksheet '{sheet_name}' was not found in {workbook_path}")

    target = None
    for relationship in relationships_xml.findall(
        f".//{{{_PACKAGE_REL_NS}}}Relationship"
    ):
        if relationship.get("Id") == relationship_id:
            target = relationship.get("Target")
            break
    if not target:
        raise ValueError(
            f"Worksheet relationship for '{sheet_name}' was not found in {workbook_path}"
        )

    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join("xl", target))


def _restore_template_extensions(
    template_path: Path,
    generated_path: Path,
    sheet_name: str,
) -> None:
    template_part = _worksheet_part_path(template_path, sheet_name)
    generated_part = _worksheet_part_path(generated_path, sheet_name)

    with ZipFile(template_path) as template_archive:
        template_xml = template_archive.read(template_part)

    template_root = _ROOT_TAG_PATTERN.search(template_xml)
    template_extensions = _EXT_LIST_PATTERN.search(template_xml)
    if template_root is None or template_extensions is None:
        return

    temporary_zip = generated_path.with_name(
        f"{generated_path.stem}.{uuid4().hex}.restore{generated_path.suffix}"
    )
    with ZipFile(generated_path) as source_archive, ZipFile(
        temporary_zip,
        mode="w",
        compression=ZIP_DEFLATED,
    ) as destination_archive:
        for item in source_archive.infolist():
            content = source_archive.read(item.filename)
            if item.filename == generated_part:
                generated_root = _ROOT_TAG_PATTERN.search(content)
                if generated_root is None:
                    raise ValueError(
                        f"Generated worksheet XML is invalid: {generated_part}"
                    )
                content = (
                    content[: generated_root.start()]
                    + template_root.group(0)
                    + content[generated_root.end() :]
                )
                content = _EXT_LIST_PATTERN.sub(b"", content)
                closing_tag = b"</worksheet>"
                closing_index = content.rfind(closing_tag)
                if closing_index < 0:
                    raise ValueError(
                        f"Generated worksheet closing tag is missing: {generated_part}"
                    )
                content = (
                    content[:closing_index]
                    + template_extensions.group(0)
                    + content[closing_index:]
                )
            destination_archive.writestr(item, content)

    os.replace(temporary_zip, generated_path)


def fill_nas_census_template(
    *,
    request_id: str,
    request_type: str,
    output_path: str | Path,
    template_path: Path,
    include_user_ids: Iterable[str] | None,
    value_mapper: ValueMapper,
    success_log_label: str,
    logger=None,
) -> CensusBuildResult:
    if not str(request_id or "").strip():
        raise ValueError("request_id is required for NAS census generation")
    if not template_path.exists():
        raise FileNotFoundError(f"NAS census template not found: {template_path}")

    members_df = load_request_members_dataframe(
        request_id=request_id,
        portal_name="NAS",
        request_type=request_type,
        include_user_ids=include_user_ids,
        logger=logger,
    )

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Data Validation extension is not supported and will be removed",
            category=UserWarning,
        )
        workbook = load_workbook(template_path)

    sheet_name = "Sample Template"
    if sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"NAS template does not contain '{sheet_name}' sheet: {template_path}"
        )
    worksheet = workbook[sheet_name]

    header_by_column: Dict[int, str] = {}
    for column_index in range(1, worksheet.max_column + 1):
        header = normalize_header(worksheet.cell(row=1, column=column_index).value)
        if header:
            header_by_column[column_index] = header
    if not header_by_column:
        raise ValueError(f"NAS template sheet has no headers: {template_path}")

    for row_index in range(2, worksheet.max_row + 1):
        for column_index in header_by_column:
            worksheet.cell(row=row_index, column=column_index).value = None

    records = members_df.to_dict(orient="records")
    for member_index, row in enumerate(records, start=1):
        excel_row = member_index + 1
        for column_index, header in header_by_column.items():
            cell = worksheet.cell(row=excel_row, column=column_index)
            cell.value = value_mapper(row, header, member_index)
            cell.number_format = "General"

    resolved_output = Path(output_path).resolve()
    resolved_output.parent.mkdir(parents=True, exist_ok=True)
    temporary_output = resolved_output.with_name(
        f"{resolved_output.stem}.{uuid4().hex}.tmp{resolved_output.suffix}"
    )
    workbook.save(temporary_output)
    workbook.close()
    _restore_template_extensions(template_path, temporary_output, sheet_name)
    os.replace(temporary_output, resolved_output)

    if logger:
        logger.info(
            "%s | RequestId=%s | Members=%s | Output=%s",
            success_log_label,
            request_id,
            len(records),
            resolved_output,
        )

    return CensusBuildResult(
        request_id=str(request_id).strip(),
        output_path=resolved_output,
        members_count=len(records),
        template_path=template_path.resolve(),
    )
