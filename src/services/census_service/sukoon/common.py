from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
import os
from pathlib import Path
from typing import Any, Callable, Dict, Iterable
from uuid import uuid4

import pandas as pd
from openpyxl import load_workbook

from src.services.db_service.azure_db_connection import AzureSQLConnection


TABLE_NAME = "[dbo].[EndorsementRequestsMemberData]"
STATUS_TABLE = "[dbo].[EndorsementRequestStatus]"


@dataclass(frozen=True)
class CensusBuildResult:
    request_id: str
    output_path: Path
    members_count: int
    template_path: Path


def normalize_header(header: Any) -> str:
    text = str(header or "").strip().lower()
    return " ".join(text.split())


def normalize_scalar(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return format(value, "f").rstrip("0").rstrip(".") or "0"
    return str(value).strip()


def to_excel_date(value: Any) -> date | str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text_value = str(value).strip()
    if not text_value:
        return ""

    parsed = pd.to_datetime(text_value, errors="coerce", dayfirst=False)
    if pd.notna(parsed):
        return parsed.date()
    return text_value


def load_request_members_dataframe(
    request_id: str,
    portal_name: str,
    request_type: str,
    include_user_ids: Iterable[str] | None = None,
    logger=None,
) -> pd.DataFrame:
    normalized_user_ids = [str(user_id).strip() for user_id in (include_user_ids or []) if str(user_id).strip()]
    where_sql = """
RequestId = ?
  AND UPPER(PortalName) = ?
  AND UPPER(RequestType) = ?
"""
    query_params: list[Any] = [
        str(request_id).strip(),
        str(portal_name).upper(),
        str(request_type).upper(),
    ]

    if normalized_user_ids:
        placeholders = ", ".join("?" for _ in normalized_user_ids)
        where_sql += f"\n  AND UserId IN ({placeholders})"
        query_params.extend(normalized_user_ids)

    query = f"""
SELECT *
FROM {TABLE_NAME}
WHERE {where_sql}
ORDER BY Id ASC, CreatedAt ASC
"""

    with AzureSQLConnection(logger=logger) as db_connection:
        connection = db_connection.connect()
        cursor = connection.cursor()
        try:
            cursor.execute(query, query_params)
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()
        finally:
            cursor.close()

    frame = pd.DataFrame.from_records(rows, columns=columns)
    if frame.empty:
        user_filter_text = f", UserFilterCount={len(normalized_user_ids)}" if normalized_user_ids else ""
        raise ValueError(
            "No member rows found for census generation with "
            f"RequestId={request_id}, PortalName={portal_name}, RequestType={request_type}{user_filter_text}"
        )
    return frame


def load_success_user_ids(request_id: str, logger=None) -> list[str]:
    query = f"""
SELECT DISTINCT UserId
FROM {STATUS_TABLE}
WHERE RequestId = ?
  AND UPPER(ISNULL(EmailStatus, '')) = 'SUCCESS'
  AND UPPER(ISNULL(OcrStatus, '')) = 'SUCCESS'
    AND UPPER(ISNULL(ValidationStatus, '')) = 'SUCCESS'
ORDER BY UserId ASC
"""

    with AzureSQLConnection(logger=logger) as db_connection:
        connection = db_connection.connect()
        cursor = connection.cursor()
        try:
            cursor.execute(query, [str(request_id).strip()])
            rows = cursor.fetchall()
        finally:
            cursor.close()

    user_ids = [str(row[0]).strip() for row in rows if str(row[0]).strip()]
    if not user_ids:
        raise ValueError(f"No OCR/Email SUCCESS members found for request RequestId={request_id}")
    return user_ids


def fill_census_template(
    *,
    request_id: str,
    request_type: str,
    output_path: str | Path,
    template_path: Path,
    portal_name: str,
    include_user_ids: Iterable[str] | None,
    value_mapper: Callable[[Dict[str, Any], str, int], Any],
    date_headers: set[str],
    success_log_label: str,
    logger=None,
) -> CensusBuildResult:
    if not str(request_id or "").strip():
        raise ValueError("request_id is required for census generation")
    if not template_path.exists():
        raise FileNotFoundError(f"Census template not found: {template_path}")

    members_df = load_request_members_dataframe(
        request_id=request_id,
        portal_name=portal_name,
        request_type=request_type,
        include_user_ids=include_user_ids,
        logger=logger,
    )

    workbook = load_workbook(template_path)
    worksheet = workbook["Members"] if "Members" in workbook.sheetnames else workbook[workbook.sheetnames[0]]

    header_row_index = 1
    header_by_column: Dict[int, str] = {}
    for column_index in range(1, worksheet.max_column + 1):
        normalized = normalize_header(worksheet.cell(row=header_row_index, column=column_index).value)
        if normalized:
            header_by_column[column_index] = normalized

    if not header_by_column:
        raise ValueError(f"Template sheet has no headers: {template_path}")

    for row_index in range(header_row_index + 1, worksheet.max_row + 1):
        for column_index in header_by_column:
            worksheet.cell(row=row_index, column=column_index).value = None

    records = members_df.to_dict(orient="records")
    for member_index, row in enumerate(records, start=1):
        excel_row = header_row_index + member_index
        for column_index, header in header_by_column.items():
            mapped_value = value_mapper(row, header, member_index)
            cell = worksheet.cell(row=excel_row, column=column_index)
            cell.value = mapped_value
            if header in date_headers and isinstance(mapped_value, date):
                cell.number_format = "MM/DD/YYYY"

    resolved_output = Path(output_path).resolve()
    resolved_output.parent.mkdir(parents=True, exist_ok=True)
    temporary_output = resolved_output.with_name(f"{resolved_output.stem}.{uuid4().hex}.tmp{resolved_output.suffix}")
    workbook.save(temporary_output)
    os.replace(temporary_output, resolved_output)

    if logger:
        logger.info(
            f"{success_log_label} | "
            f"RequestId={request_id} | Members={len(records)} | Output={resolved_output}"
        )

    return CensusBuildResult(
        request_id=str(request_id).strip(),
        output_path=resolved_output,
        members_count=len(records),
        template_path=template_path.resolve(),
    )
