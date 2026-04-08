from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook

from src.services.db_service.azure_db_connection import AzureSQLConnection


TABLE_NAME = "[dbo].[EndorsementRequestsMemberData]"
DEFAULT_ADD_TEMPLATE = Path("data/templates/SUKOON/ADD.xlsx")
DEFAULT_DELETE_TEMPLATE = Path("data/templates/SUKOON/DELETE.xlsx")


@dataclass(frozen=True)
class CensusBuildResult:
    request_id: str
    output_path: Path
    members_count: int
    template_path: Path


def _normalize_header(header: Any) -> str:
    text = str(header or "").strip().lower()
    text = " ".join(text.split())
    return text


def _normalize_scalar(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return format(value, "f").rstrip("0").rstrip(".") or "0"
    return str(value).strip()


def _to_excel_date(value: Any) -> date | str:
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text_value = str(value).strip()
    if not text_value:
        return ""

    parsed = pd.to_datetime(text_value, errors="coerce", dayfirst=False)
    if pd.notna(parsed):
        return parsed.date()

    return text_value


def _value_from_row(row: Dict[str, Any], header: str, row_number: int) -> Any:
    if header == "sl no.":
        return row_number
    if header == "first name":
        return _normalize_scalar(row.get("FirstName"))
    if header == "middle name":
        return _normalize_scalar(row.get("MiddleName"))
    if header == "last name":
        return _normalize_scalar(row.get("LastName"))
    if header == "employee number":
        return _normalize_scalar(row.get("StaffId") or row.get("HealthCardNumber"))
    if header == "date of birth":
        return _to_excel_date(row.get("DateOfBirth"))
    if header == "gender":
        return _normalize_scalar(row.get("Gender"))
    if header == "marital status":
        return _normalize_scalar(row.get("MaritalStatus"))
    if header == "relation":
        return _normalize_scalar(row.get("Relation"))
    if header == "category":
        return _normalize_scalar(row.get("Category"))
    if header == "region":
        return _normalize_scalar(row.get("WorkRegion") or row.get("ResidenceRegion"))
    if header == "lsb":
        return _normalize_scalar(row.get("SalaryBand"))
    if header == "nationality":
        return _normalize_scalar(row.get("Nationality"))
    if header == "passport number":
        return _normalize_scalar(row.get("PassportNo"))
    if header == "eid number":
        return _normalize_scalar(row.get("EmiratesId"))
    if header == "uid number":
        return _normalize_scalar(row.get("UnifiedNo"))
    if header == "visa issued location":
        return _normalize_scalar(row.get("VisaIssuanceEmirate"))
    if header == "actual salary band":
        return _normalize_scalar(row.get("SalaryBand"))
    if header == "person commission":
        return _normalize_scalar(row.get("Commission"))
    if header == "residential location":
        return _normalize_scalar(row.get("ResidenceEmirate") or row.get("ResidenceRegion"))
    if header == "work location":
        return _normalize_scalar(row.get("WorkEmirate") or row.get("WorkRegion"))
    if header == "mobile number":
        return _normalize_scalar(row.get("MobileNo"))
    if header == "email":
        return _normalize_scalar(row.get("Email"))
    if header == "photo file name":
        return _normalize_scalar(row.get("MemberPhoto"))
    if header == "sponsor type":
        return _normalize_scalar(row.get("SponsorType"))
    if header == "sponsor id":
        return _normalize_scalar(row.get("SponsorId"))
    if header == "sponsor contact number":
        return _normalize_scalar(row.get("SponsorContactNumber"))
    if header == "sponsor contact email":
        return _normalize_scalar(row.get("SponsorContactEmail"))
    if header == "occupation":
        return _normalize_scalar(row.get("Occupation"))
    if header in {"addition effective date", "effective date", "deletion effective date"}:
        return _to_excel_date(row.get("DeletionEffectiveDate") or row.get("EffectiveDate"))
    if header == "visa file number":
        return _normalize_scalar(row.get("VisaFileNumber"))
    if header == "birth certificate number":
        return _normalize_scalar(row.get("BirthCertificateNumber"))
    if header == "policy number":
        return _normalize_scalar(row.get("PolicyNumber"))
    if header in {"health card number", "card number"}:
        return _normalize_scalar(row.get("HealthCardNumber") or row.get("StaffId"))

    return ""


def _load_request_members_dataframe(
    request_id: str,
    portal_name: str,
    request_type: str,
    include_user_ids: Iterable[str] | None = None,
    logger=None,
) -> pd.DataFrame:
    normalized_user_ids = [str(user_id).strip() for user_id in (include_user_ids or []) if str(user_id).strip()]
    where_sql = """
RequestId = ?
  AND UPPER(PortalName) = ?
  AND UPPER(RequestType) = ?
"""
    query_params: list[Any] = [
        str(request_id).strip(),
        str(portal_name).upper(),
        str(request_type).upper(),
    ]

    if normalized_user_ids:
        placeholders = ", ".join("?" for _ in normalized_user_ids)
        where_sql += f"\n  AND UserId IN ({placeholders})"
        query_params.extend(normalized_user_ids)

    query = f"""
SELECT *
FROM {TABLE_NAME}
WHERE {where_sql}
ORDER BY Id ASC, CreatedAt ASC
"""

    with AzureSQLConnection(logger=logger) as db_connection:
        connection = db_connection.connect()
        cursor = connection.cursor()
        try:
            cursor.execute(query, query_params)
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()
        finally:
            cursor.close()

    frame = pd.DataFrame.from_records(rows, columns=columns)

    if frame.empty:
        user_filter_text = ""
        if normalized_user_ids:
            user_filter_text = f", UserFilterCount={len(normalized_user_ids)}"
        raise ValueError(
            "No member rows found for census generation with "
            f"RequestId={request_id}, PortalName={portal_name}, RequestType={request_type}{user_filter_text}"
        )

    return frame


def _resolve_template_path(request_type: str) -> Path:
    if str(request_type).strip().upper() == "DELETE" and DEFAULT_DELETE_TEMPLATE.exists():
        return DEFAULT_DELETE_TEMPLATE
    return DEFAULT_ADD_TEMPLATE


def build_batch_census_file(
    request_id: str,
    request_type: str,
    output_path: str | Path,
    portal_name: str = "SUKOON",
    include_user_ids: Iterable[str] | None = None,
    logger=None,
) -> CensusBuildResult:
    if not str(request_id or "").strip():
        raise ValueError("request_id is required for batch census generation")

    template_path = _resolve_template_path(request_type=request_type)
    if not template_path.exists():
        raise FileNotFoundError(f"Census template not found: {template_path}")

    members_df = _load_request_members_dataframe(
        request_id=request_id,
        portal_name=portal_name,
        request_type=request_type,
        include_user_ids=include_user_ids,
        logger=logger,
    )

    workbook = load_workbook(template_path)
    worksheet = workbook["Members"] if "Members" in workbook.sheetnames else workbook[workbook.sheetnames[0]]

    header_row_index = 1
    header_by_column: Dict[int, str] = {}
    for column_index in range(1, worksheet.max_column + 1):
        cell_value = worksheet.cell(row=header_row_index, column=column_index).value
        normalized = _normalize_header(cell_value)
        if normalized:
            header_by_column[column_index] = normalized

    if not header_by_column:
        raise ValueError(f"Template sheet has no headers: {template_path}")

    date_headers = {
        "date of birth",
        "addition effective date",
        "effective date",
        "deletion effective date",
    }

    for row_index in range(header_row_index + 1, worksheet.max_row + 1):
        for column_index in header_by_column:
            worksheet.cell(row=row_index, column=column_index).value = None

    records = members_df.to_dict(orient="records")
    for member_index, row in enumerate(records, start=1):
        excel_row = header_row_index + member_index
        for column_index, header in header_by_column.items():
            mapped_value = _value_from_row(
                row=row,
                header=header,
                row_number=member_index,
            )
            cell = worksheet.cell(row=excel_row, column=column_index)
            cell.value = mapped_value
            if header in date_headers and isinstance(mapped_value, date):
                cell.number_format = "MM/DD/YYYY"

    resolved_output = Path(output_path).resolve()
    resolved_output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(resolved_output)

    if logger:
        logger.info(
            "Batch census generated | "
            f"RequestId={request_id} | Members={len(records)} | Output={resolved_output}"
        )

    return CensusBuildResult(
        request_id=str(request_id).strip(),
        output_path=resolved_output,
        members_count=len(records),
        template_path=template_path.resolve(),
    )
